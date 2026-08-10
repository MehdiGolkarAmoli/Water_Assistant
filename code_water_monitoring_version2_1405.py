"""
Water Quality Monitoring Application
=====================================
Sentinel-2 based automatic calculation of:
  - Water Turbidity (NDTI)
  - Chlorophyll-a Concentration

Both indices share the same preprocessing pipeline (cloud filtering, snow
masking, waterbody extraction) and are now computed automatically and
sequentially after the user selects an area of interest. The interface is
designed for managers and non-technical decision-makers: no remote-sensing
jargon, no parameter pickers, no processing logs.
"""

import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from shapely.geometry import Polygon
import rasterio
import datetime
import math
import ee
import tempfile
import requests
import time
import warnings
import base64
import json
from datetime import date
from PIL import Image

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)

import streamlit as st

st.set_page_config(
    layout="wide",
    page_title="Water Quality Monitoring",
    page_icon="🌊"
)

import folium
from folium import plugins
from streamlit_folium import st_folium

# =============================================================================
# CONSTANTS
# =============================================================================
# --- Fixed processing thresholds (no longer user-configurable) -------------
CLOUD_THRESHOLD = 10          # % — fixed cloud coverage threshold (CLOUDY_PIXEL_PERCENTAGE)
CLOUD_PROB_THRESHOLD = 15      # per-pixel cloud probability cutoff

# Water body detection threshold
# AWEIsh (Automated Water Extraction Index, shadow variant) — shared by both
# indices so NDTI and NDCI are computed on the exact same water mask.
# AWEIsh = Blue + 2.5*Green - 1.5*(NIR + SWIR1) - 0.25*SWIR2
AWEI_THRESHOLD = 0.05

# Snow detection thresholds (preprocessing only — excludes snow from water)
NDSI_THRESHOLD = 0.39
SNOW_B11_THRESHOLD = 0.1  # kept for reference; no longer used by is_snow (see below)

# MODIS-heritage water/snow discrimination test (Hall et al., 1995; Riggs et al.),
# translated to Sentinel-2 bands. Water absorbs NIR almost completely regardless
# of turbidity, while snow reflects strongly there, so this is a more robust way
# to keep turbid/sediment-laden water from being flagged as snow than the SWIR
# threshold alone.
NIR_SNOW_THRESHOLD = 0.11    # B8 (NIR) — MODIS band 2 analogue
GREEN_SNOW_THRESHOLD = 0.1   # B3 (Green) — MODIS band 4 analogue

# Parameter identifiers (internal use only — never shown as a user choice)
PARAM_TURBIDITY = "Turbidity (NDTI)"
PARAM_CHLOROPHYLL = "Chlorophyll Index"

# Chlorophyll visualization range
CHL_VMIN = -1.0
CHL_VMAX = 0.9

# Download settings
MAX_RETRIES = 3
RETRY_DELAY_BASE = 2
DOWNLOAD_TIMEOUT = 120
CHUNK_SIZE = 8192
MIN_FILE_SIZE = 10000

# Status constants
STATUS_NO_DATA = "no_data"
STATUS_COMPLETE = "complete"
STATUS_FAILED = "failed"
STATUS_SKIPPED = "skipped"

# =============================================================================
# Session State Initialization
# =============================================================================
if 'drawn_polygons' not in st.session_state:
    st.session_state.drawn_polygons = []
if 'last_drawn_polygon' not in st.session_state:
    st.session_state.last_drawn_polygon = None
if 'ee_initialized' not in st.session_state:
    st.session_state.ee_initialized = False
if 'current_temp_dir' not in st.session_state:
    st.session_state.current_temp_dir = None
if 'downloaded_months' not in st.session_state:
    # nested by parameter: {PARAM_TURBIDITY: {...}, PARAM_CHLOROPHYLL: {...}}
    st.session_state.downloaded_months = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
if 'month_statuses' not in st.session_state:
    st.session_state.month_statuses = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
if 'results' not in st.session_state:
    # nested by parameter
    st.session_state.results = {PARAM_TURBIDITY: [], PARAM_CHLOROPHYLL: []}
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'selected_region_index' not in st.session_state:
    st.session_state.selected_region_index = 0
if 'processing_in_progress' not in st.session_state:
    st.session_state.processing_in_progress = False
if 'processing_config' not in st.session_state:
    st.session_state.processing_config = None
if 'mean_data' not in st.session_state:
    st.session_state.mean_data = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
if 'download_summary' not in st.session_state:
    # simple end-user facing summary: {PARAM_TURBIDITY: (downloaded, available), ...}
    st.session_state.download_summary = {}
if 'resume_after_interruption' not in st.session_state:
    # True when a previous run was interrupted and can be resumed
    st.session_state.resume_after_interruption = False


@st.cache_data(ttl=None, persist="disk", show_spinner=False)
def _fetch_monthly_weather_cached(lat: float, lon: float, year: int, month: int) -> dict:
    last_day = calendar.monthrange(year, month)[1]
    params = {
        "latitude": round(lat, 3), "longitude": round(lon, 3),
        "start_date": f"{year}-{month:02d}-01",
        "end_date": f"{year}-{month:02d}-{last_day}",
        "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,snowfall_sum,wind_speed_10m_max",
        "timezone": "auto",
    }
    r = requests.get("https://archive-api.open-meteo.com/v1/archive",
                      params=params, timeout=(5, 15))  # (connect, read) — fail fast, not at 30s
    r.raise_for_status()

# =============================================================================
# Earth Engine Authentication
# =============================================================================
@st.cache_resource
def initialize_earth_engine():
    """Initialize Earth Engine"""
    try:
        ee.Initialize()
        return True, "Earth Engine initialized"
    except Exception:
        try:
            base64_key = os.environ.get('GOOGLE_EARTH_ENGINE_KEY_BASE64')

            if base64_key:
                key_json = base64.b64decode(base64_key).decode()
                key_data = json.loads(key_json)

                key_file = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
                with open(key_file.name, 'w') as f:
                    json.dump(key_data, f)

                credentials = ee.ServiceAccountCredentials(key_data['client_email'], key_file.name)
                ee.Initialize(credentials)
                os.unlink(key_file.name)
                return True, "Authenticated with Service Account"
            else:
                ee.Authenticate()
                ee.Initialize()
                return True, "Authenticated"
        except Exception as auth_error:
            return False, f"Auth failed: {str(auth_error)}"


# =============================================================================
# Helper Functions
# =============================================================================
def get_utm_zone(longitude):
    return math.floor((longitude + 180) / 6) + 1


def validate_geotiff_file(file_path, expected_bands=1):
    """Validate that a GeoTIFF file is complete and readable."""
    try:
        if not os.path.exists(file_path):
            return False, "File does not exist"

        file_size = os.path.getsize(file_path)
        if file_size < MIN_FILE_SIZE:
            return False, f"File too small ({file_size} bytes)"

        with rasterio.open(file_path) as src:
            if src.count < expected_bands:
                return False, f"Wrong band count ({src.count}, expected {expected_bands})"

        return True, "File is valid"

    except Exception as e:
        return False, f"Validation error: {str(e)}"


# =============================================================================
# Water Quality Calculation (GEE Server-Side)
# =============================================================================
def create_water_quality_collection(aoi, start_date, end_date, parameter_type, cloudy_pixel_percentage=CLOUD_THRESHOLD):
    """
    Create water quality collection for either Turbidity or Chlorophyll.

    Snow detection is used as a PREPROCESSING step to exclude snow/ice pixels
    from water detection. Snow mask is never downloaded or shown to the user.

    For TURBIDITY (NDTI):
    1. Link S2_SR with S2_CLOUD_PROBABILITY
    2. Apply cloud mask (probability < 15)
    3. Calculate NDSI for snow detection: (B3 - B11) / (B3 + B11)
    4. Create snow mask (MODIS-heritage water/snow test): NDSI > 0.42 AND B8 (NIR) > 0.11 AND B3 (Green) > 0.1
    5. Calculate AWEIsh for water body detection: B2 + 2.5*B3 - 1.5*(B8+B11) - 0.25*B12 > 0.05, excluding snow
    6. Calculate NDTI (turbidity index): (B4 - B3) / (B4 + B3)

    For CHLOROPHYLL:
    1. Link S2_SR with S2_CLOUD_PROBABILITY
    2. Apply cloud mask (probability < 15)
    3. Calculate NDSI for snow detection: (B3 - B11) / (B3 + B11)
    4. Create snow mask (MODIS-heritage water/snow test): NDSI > 0.42 AND B8 (NIR) > 0.11 AND B3 (Green) > 0.1
    5. Calculate AWEIsh for water body detection: B2 + 2.5*B3 - 1.5*(B8+B11) - 0.25*B12 > 0.05, excluding snow
    6. Calculate Chlorophyll Index (NDCI): (B5 - B4) / (B5 + B4)
    """
    s2_sr = (ee.ImageCollection('COPERNICUS/S2_SR_HARMONIZED')
             .filterBounds(aoi)
             .filterDate(start_date, end_date)
             .filter(ee.Filter.lt('CLOUDY_PIXEL_PERCENTAGE', cloudy_pixel_percentage)))

    s2_cloud_prob = (ee.ImageCollection('COPERNICUS/S2_CLOUD_PROBABILITY')
                     .filterBounds(aoi)
                     .filterDate(start_date, end_date))

    join_filter = ee.Filter.equals(leftField='system:index', rightField='system:index')
    joined = ee.Join.saveFirst('cloud_probability').apply(
        primary=s2_sr, secondary=s2_cloud_prob, condition=join_filter
    )

    def add_cloud_band(feature):
        img = ee.Image(feature)
        cloud_prob_img = ee.Image(img.get('cloud_probability'))
        return img.addBands(cloud_prob_img.select('probability'))

    s2_joined = ee.ImageCollection(joined.map(add_cloud_band))

    if parameter_type == PARAM_TURBIDITY:
        def calculate_turbidity(img):
            cloud = img.select('probability')
            cloud_free = cloud.lt(CLOUD_PROB_THRESHOLD)

            sr = img.select(['B2', 'B3', 'B4', 'B8', 'B11', 'B12']).multiply(0.0001)

            ndsi = sr.normalizedDifference(['B3', 'B11']).rename('ndsi')
            is_snow = (
                ndsi.gt(NDSI_THRESHOLD)                          # NDSI > 0.42 — spectral snow signature
                .And(sr.select('B8').gt(NIR_SNOW_THRESHOLD))     # NIR ~0.11 — excludes water, snow reflects strongly here
                .And(sr.select('B3').gt(GREEN_SNOW_THRESHOLD))   # Green ~0.1 — excludes dark shadow/non-snow surfaces
            )

            awei = sr.expression(
                'BLUE + 2.5 * GREEN - 1.5 * (NIR + SWIR1) - 0.25 * SWIR2',
                {
                    'BLUE': sr.select('B2'),
                    'GREEN': sr.select('B3'),
                    'NIR': sr.select('B8'),
                    'SWIR1': sr.select('B11'),
                    'SWIR2': sr.select('B12'),
                }
            ).rename('awei')
            water_body = awei.gt(AWEI_THRESHOLD).And(is_snow.Not())

            ndti = sr.normalizedDifference(['B4', 'B3']).rename('wq_index')

            wq_masked = ndti.updateMask(cloud_free).updateMask(water_body)

            rgb = sr.select(['B4', 'B3', 'B2'])

            combined = (wq_masked
                       .addBands(rgb)
                       .addBands(water_body.rename('water_mask')))

            return combined.clip(aoi).copyProperties(img, ['system:time_start'])

        return s2_joined.map(calculate_turbidity)

    else:  # CHLOROPHYLL
        def calculate_chlorophyll(img):
            cloud = img.select('probability')
            cloud_free = cloud.lt(CLOUD_PROB_THRESHOLD)

            sr = img.select(['B1', 'B2', 'B3', 'B4', 'B5', 'B8', 'B11', 'B12']).multiply(0.0001)

            ndsi = sr.normalizedDifference(['B3', 'B11']).rename('ndsi')
            is_snow = (
                ndsi.gt(NDSI_THRESHOLD)                          # NDSI > 0.42 — spectral snow signature
                .And(sr.select('B8').gt(NIR_SNOW_THRESHOLD))     # NIR ~0.11 — excludes water, snow reflects strongly here
                .And(sr.select('B3').gt(GREEN_SNOW_THRESHOLD))   # Green ~0.1 — excludes dark shadow/non-snow surfaces
            )

            awei = sr.expression(
                'BLUE + 2.5 * GREEN - 1.5 * (NIR + SWIR1) - 0.25 * SWIR2',
                {
                    'BLUE': sr.select('B2'),
                    'GREEN': sr.select('B3'),
                    'NIR': sr.select('B8'),
                    'SWIR1': sr.select('B11'),
                    'SWIR2': sr.select('B12'),
                }
            ).rename('awei')
            water_body = awei.gt(AWEI_THRESHOLD).And(is_snow.Not())

            # NDCI (Normalized Difference Chlorophyll Index): (B5 - B4) / (B5 + B4)
            # https://custom-scripts.sentinel-hub.com/custom-scripts/sentinel-2/ndci/
            chl_index = sr.normalizedDifference(['B5', 'B4']).rename('wq_index')

            wq_masked = chl_index.updateMask(cloud_free).updateMask(water_body)

            rgb = sr.select(['B4', 'B3', 'B2'])

            combined = (wq_masked
                       .addBands(rgb)
                       .addBands(water_body.rename('water_mask')))

            return combined.clip(aoi).copyProperties(img, ['system:time_start'])

        return s2_joined.map(calculate_chlorophyll)


def get_monthly_composite(wq_collection, aoi, year, month):
    """Create monthly composite from water quality collection."""
    start = ee.Date.fromYMD(year, month, 1)
    end = start.advance(1, 'month')

    monthly = wq_collection.filterDate(start, end)
    count = monthly.size().getInfo()

    if count == 0:
        return None, 0, "No images"

    composite = monthly.median()

    stats = composite.select('wq_index').reduceRegion(
        reducer=ee.Reducer.mean().combine(
            ee.Reducer.count(), sharedInputs=True
        ).combine(
            ee.Reducer.minMax(), sharedInputs=True
        ),
        geometry=aoi,
        scale=10,
        maxPixels=1e13
    )

    return composite, count, stats


# =============================================================================
# Download Functions
# =============================================================================
def download_band_with_retry(image, band, aoi, output_path, scale=10):
    """Download a single band with retry mechanism."""
    try:
        region = aoi.bounds().getInfo()['coordinates']
    except Exception as e:
        return False, f"AOI bounds error: {e}"

    temp_path = output_path + '.tmp'
    if os.path.exists(temp_path):
        os.remove(temp_path)

    if os.path.exists(output_path):
        is_valid, msg = validate_geotiff_file(output_path, expected_bands=1)
        if is_valid:
            return True, "cached"
        os.remove(output_path)

    last_error = None

    for attempt in range(MAX_RETRIES):
        try:
            url = image.select(band).getDownloadURL({
                'scale': scale, 'region': region, 'format': 'GEO_TIFF', 'bands': [band]
            })

            response = requests.get(url, stream=True, timeout=DOWNLOAD_TIMEOUT)

            if response.status_code == 200:
                content_type = response.headers.get('content-type', '')
                if 'text/html' in content_type:
                    last_error = "GEE rate limit"
                    raise Exception(last_error)

                downloaded_size = 0
                with open(temp_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=CHUNK_SIZE):
                        if chunk:
                            f.write(chunk)
                            downloaded_size += len(chunk)

                if downloaded_size < MIN_FILE_SIZE:
                    last_error = f"File too small ({downloaded_size} bytes)"
                    raise Exception(last_error)

                is_valid, msg = validate_geotiff_file(temp_path, expected_bands=1)
                if is_valid:
                    os.replace(temp_path, output_path)
                    return True, "success"
                else:
                    last_error = f"Validation failed: {msg}"
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    raise Exception(last_error)
            else:
                last_error = f"HTTP {response.status_code}"
                raise Exception(last_error)

        except requests.exceptions.Timeout:
            last_error = "Timeout"
        except requests.exceptions.ConnectionError:
            last_error = "Connection error"
        except Exception as e:
            if last_error is None:
                last_error = str(e)

        for f in [output_path, temp_path]:
            if os.path.exists(f):
                try:
                    os.remove(f)
                except:
                    pass

        if attempt < MAX_RETRIES - 1:
            wait_time = RETRY_DELAY_BASE ** (attempt + 1)
            time.sleep(wait_time)

    return False, last_error


def download_monthly_data(composite, aoi, temp_dir, month_name, param_short, scale=10):
    """
    Download monthly composite (Water Quality Index + RGB bands).
    Snow mask is NOT downloaded — used only server-side in GEE.
    No UI status is written here; progress is summarized at a higher level.
    """
    wq_path = os.path.join(temp_dir, f"wq_index_{param_short}_{month_name}.tif")
    rgb_path = os.path.join(temp_dir, f"rgb_{param_short}_{month_name}.tif")

    wq_valid, _ = validate_geotiff_file(wq_path, expected_bands=1)
    rgb_valid, _ = validate_geotiff_file(rgb_path, expected_bands=3)

    if wq_valid and rgb_valid:
        return wq_path, rgb_path, STATUS_COMPLETE, "Cached"

    try:
        success, msg = download_band_with_retry(composite, 'wq_index', aoi, wq_path, scale)
        if not success:
            return None, None, STATUS_FAILED, f"WQ Index download failed: {msg}"

        bands_dir = os.path.join(temp_dir, f"bands_{param_short}_{month_name}")
        os.makedirs(bands_dir, exist_ok=True)

        rgb_bands = ['B4', 'B3', 'B2']
        band_files = []

        for band in rgb_bands:
            band_file = os.path.join(bands_dir, f"{band}.tif")
            success, msg = download_band_with_retry(composite, band, aoi, band_file, scale)

            if not success:
                return None, None, STATUS_FAILED, f"RGB {band} download failed: {msg}"

            band_files.append(band_file)

        with rasterio.open(band_files[0]) as src:
            meta = src.meta.copy()
        meta.update(count=3)

        with rasterio.open(rgb_path, 'w', **meta) as dst:
            for i, band_file in enumerate(band_files):
                with rasterio.open(band_file) as src:
                    dst.write(src.read(1), i+1)

        return wq_path, rgb_path, STATUS_COMPLETE, "Downloaded"

    except Exception as e:
        return None, None, STATUS_FAILED, f"Error: {str(e)}"


# =============================================================================
# Visualization Functions
# =============================================================================
def create_turbidity_colormap():
    colors = ['#0000FF', '#00FFFF', '#00FF00', '#FFFF00', '#FF8000', '#FF0000']
    return LinearSegmentedColormap.from_list('turbidity', colors, N=256)


def create_chlorophyll_colormap():
    colors = ['#9400D3', '#4B0082', '#0000FF', '#00FF00', '#FFFF00', '#FF7F00', '#FF0000']
    return LinearSegmentedColormap.from_list('chlorophyll', colors, N=256)


def generate_thumbnails(wq_path, rgb_path, month_name, parameter_type, max_size=300):
    """Generate RGB and water quality index thumbnails."""
    try:
        with rasterio.open(wq_path) as src:
            wq_data = src.read(1)

        with rasterio.open(rgb_path) as src:
            red = src.read(1)
            green = src.read(2)
            blue = src.read(3)

        rgb = np.stack([red, green, blue], axis=-1)
        rgb = np.nan_to_num(rgb, nan=0.0)

        def percentile_stretch(band, lower=2, upper=98):
            valid = band[band > 0]
            if len(valid) == 0:
                return np.zeros_like(band, dtype=np.uint8)
            p_low = np.percentile(valid, lower)
            p_high = np.percentile(valid, upper)
            if p_high <= p_low:
                p_high = p_low + 0.001
            stretched = np.clip((band - p_low) / (p_high - p_low), 0, 1)
            return (stretched * 255).astype(np.uint8)

        rgb_uint8 = np.zeros_like(rgb, dtype=np.uint8)
        for i in range(3):
            rgb_uint8[:, :, i] = percentile_stretch(rgb[:, :, i])

        wq_valid = np.nan_to_num(wq_data, nan=np.nan)

        valid_wq = wq_valid[~np.isnan(wq_valid) & (wq_valid != 0)]
        mean_value = np.nanmean(valid_wq) if len(valid_wq) > 0 else np.nan
        valid_pixel_count = len(valid_wq)
        total_pixels = wq_data.size
        water_coverage = (valid_pixel_count / total_pixels) * 100 if total_pixels > 0 else 0

        if parameter_type == PARAM_TURBIDITY:
            cmap = create_turbidity_colormap()
            wq_normalized = np.clip((wq_valid + 0.3) / 0.6, 0, 1)
        else:
            cmap = create_chlorophyll_colormap()
            wq_normalized = np.clip((wq_valid - CHL_VMIN) / (CHL_VMAX - CHL_VMIN), 0, 1)

        wq_normalized = np.nan_to_num(wq_normalized, nan=0)

        wq_colored = cmap(wq_normalized)[:, :, :3]
        wq_uint8 = (wq_colored * 255).astype(np.uint8)

        water_mask = (~np.isnan(wq_valid)) & (wq_valid != 0)
        for i in range(3):
            wq_uint8[:, :, i] = np.where(water_mask, wq_uint8[:, :, i], 50)

        pil_rgb = Image.fromarray(rgb_uint8, mode='RGB')
        pil_wq = Image.fromarray(wq_uint8, mode='RGB')

        h, w = pil_rgb.size[1], pil_rgb.size[0]
        if h > max_size or w > max_size:
            scale = max_size / max(h, w)
            new_w, new_h = int(w * scale), int(h * scale)
            pil_rgb = pil_rgb.resize((new_w, new_h), Image.LANCZOS)
            pil_wq = pil_wq.resize((new_w, new_h), Image.LANCZOS)

        return {
            'rgb_image': pil_rgb,
            'wq_image': pil_wq,
            'month_name': month_name,
            'mean_value': mean_value,
            'water_coverage': water_coverage,
            'valid_pixels': valid_pixel_count,
            'parameter_type': parameter_type
        }

    except Exception:
        # Quietly skip a month that fails to render rather than surfacing
        # remote-sensing error internals to a non-technical user.
        return None


# =============================================================================
# Main Processing Pipeline (silent — no remote-sensing internals shown)
# =============================================================================
def process_single_parameter(aoi, start_date, end_date, parameter_type, temp_dir,
                              cloudy_pixel_percentage=CLOUD_THRESHOLD, scale=10,
                              resume=False, progress_callback=None):
    """
    Run the full pipeline for one parameter (NDTI or Chlorophyll-a):
    cloud filtering -> snow masking -> waterbody extraction -> index calculation
    -> download -> thumbnail generation.

    Key resilience behaviours (adapted from old version):
    - Per-month session state writes: progress is preserved after every month so
      that a connection drop never discards completed work.
    - Resume logic: when resume=True, months already present in
      st.session_state.downloaded_months[parameter_type] (and whose files are
      still valid on disk) are skipped entirely.
    - File-level cache: download_monthly_data() validates existing GeoTIFFs on
      disk before attempting a new download — so cached files survive page
      reloads even without session state.
    - EE server calls (.size().getInfo(), get_monthly_composite) are wrapped in
      try/except so a transient network error on one month does not crash the
      whole pipeline; the month is marked STATUS_FAILED and processing continues.

    Returns: (results_list, mean_data_dict, downloaded_count, available_count)
    No technical logs are written to the UI; this function is silent.

    progress_callback(done, total, month_name), if given, is invoked once
    before the download loop starts (to reflect months already recovered
    from cache/resume) and once after every month is handled, so the caller
    can drive a progress bar / stage label.
    """
    param_short = "turbidity" if parameter_type == PARAM_TURBIDITY else "chlorophyll"

    start_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    total_months = (end_dt.year - start_dt.year) * 12 + (end_dt.month - start_dt.month)

    # ------------------------------------------------------------------
    # Build the GEE collection (server-side — no network download yet)
    # ------------------------------------------------------------------
    wq_collection = create_water_quality_collection(
        aoi, start_date, end_date, parameter_type, cloudy_pixel_percentage
    )

    month_infos = []
    for month_index in range(total_months):
        year = start_dt.year + (start_dt.month - 1 + month_index) // 12
        month = (start_dt.month - 1 + month_index) % 12 + 1
        month_infos.append({'month_name': f"{year}-{month:02d}", 'year': year, 'month': month})

    # ------------------------------------------------------------------
    # FIX A: Restore already-downloaded months from session state (resume)
    # ------------------------------------------------------------------
    # Ensure the nested dict for this parameter exists in session state
    if not isinstance(st.session_state.downloaded_months.get(parameter_type), dict):
        st.session_state.downloaded_months[parameter_type] = {}
    if not isinstance(st.session_state.month_statuses.get(parameter_type), dict):
        st.session_state.month_statuses[parameter_type] = {}

    downloaded_months = {}

    if resume and st.session_state.downloaded_months.get(parameter_type):
        for month_name, paths in st.session_state.downloaded_months[parameter_type].items():
            if paths.get('wq_index') and paths.get('rgb'):
                wq_valid, _ = validate_geotiff_file(paths['wq_index'], expected_bands=1)
                rgb_valid, _ = validate_geotiff_file(paths['rgb'], expected_bands=3)
                if wq_valid and rgb_valid:
                    downloaded_months[month_name] = paths
                    # Keep the cached status entry as well
                    if month_name not in st.session_state.month_statuses[parameter_type]:
                        st.session_state.month_statuses[parameter_type][month_name] = {
                            'status': STATUS_COMPLETE, 'message': 'Cached'
                        }

    # Also check disk for any months whose files exist but session state was lost
    # (e.g. after a page reload without resume — file-level cache recovery)
    for month_info in month_infos:
        month_name = month_info['month_name']
        if month_name in downloaded_months:
            continue
        wq_path = os.path.join(temp_dir, f"wq_index_{param_short}_{month_name}.tif")
        rgb_path = os.path.join(temp_dir, f"rgb_{param_short}_{month_name}.tif")
        wq_valid, _ = validate_geotiff_file(wq_path, expected_bands=1)
        rgb_valid, _ = validate_geotiff_file(rgb_path, expected_bands=3)
        if wq_valid and rgb_valid:
            downloaded_months[month_name] = {'wq_index': wq_path, 'rgb': rgb_path}
            st.session_state.downloaded_months[parameter_type][month_name] = downloaded_months[month_name]
            st.session_state.month_statuses[parameter_type][month_name] = {
                'status': STATUS_COMPLETE, 'message': 'Cached (disk)'
            }

    # Months not yet downloaded (skip ones already done or already statused as no-data)
    already_statused = {
        m for m, s in st.session_state.month_statuses[parameter_type].items()
        if s.get('status') in (STATUS_NO_DATA, STATUS_COMPLETE)
    }
    months_to_process = [
        m for m in month_infos
        if m['month_name'] not in downloaded_months
        and m['month_name'] not in already_statused
    ]

    available_count = len(downloaded_months)  # start with already-recovered months

    # Progress bookkeeping: months already resolved before the loop (from the
    # resume/disk-cache recovery above) count as "done" immediately so the
    # progress bar reflects real state instead of restarting from zero.
    already_done_count = total_months - len(months_to_process)
    processed_count = already_done_count
    if progress_callback:
        progress_callback(processed_count, total_months, None)

    # ------------------------------------------------------------------
    # FIX B: Per-month EE + download loop with immediate session state writes
    # ------------------------------------------------------------------
    for month_info in months_to_process:
        month_name = month_info['month_name']

        # FIX C: Wrap every EE server call so a transient error skips the month
        try:
            composite, count, stats = get_monthly_composite(
                wq_collection, aoi, month_info['year'], month_info['month']
            )
        except Exception:
            # Network or EE error — mark as failed and continue to next month
            st.session_state.month_statuses[parameter_type][month_name] = {
                'status': STATUS_FAILED, 'message': 'EE request failed'
            }
            processed_count += 1
            if progress_callback:
                progress_callback(processed_count, total_months, month_name)
            continue

        if composite is None or count == 0:
            st.session_state.month_statuses[parameter_type][month_name] = {
                'status': STATUS_NO_DATA, 'message': 'No images'
            }
            processed_count += 1
            if progress_callback:
                progress_callback(processed_count, total_months, month_name)
            continue

        available_count += 1

        wq_path, rgb_path, status, message = download_monthly_data(
            composite, aoi, temp_dir, month_name, param_short, scale
        )

        # FIX D: Write to session state immediately after each month — not at end
        st.session_state.month_statuses[parameter_type][month_name] = {
            'status': status, 'message': message
        }

        if status == STATUS_COMPLETE:
            downloaded_months[month_name] = {'wq_index': wq_path, 'rgb': rgb_path}
            # Persist to session state right away so a crash/reload can recover
            st.session_state.downloaded_months[parameter_type][month_name] = {
                'wq_index': wq_path, 'rgb': rgb_path
            }

        processed_count += 1
        if progress_callback:
            progress_callback(processed_count, total_months, month_name)

    # ------------------------------------------------------------------
    # Thumbnail generation (uses only successfully downloaded months)
    # ------------------------------------------------------------------
    results = []
    mean_data = {}

    for month_name in sorted(downloaded_months.keys()):
        paths = downloaded_months[month_name]
        thumb = generate_thumbnails(paths['wq_index'], paths['rgb'], month_name, parameter_type)
        if thumb:
            results.append(thumb)
            mean_data[month_name] = {'mean': thumb['mean_value'], 'coverage': thumb['water_coverage']}

    return results, mean_data, len(downloaded_months), available_count


def run_full_analysis(aoi, start_date, end_date, cloudy_pixel_percentage=CLOUD_THRESHOLD,
                       scale=10, resume=False):
    """
    Automatically runs preprocessing + both index calculations (NDTI, then
    Chlorophyll-a) in sequence. Displays only a simple, user-friendly summary.

    Resilience additions vs. original:
    - resume=True is forwarded to process_single_parameter so cached months are
      skipped rather than re-downloaded.
    - Each parameter block is wrapped in try/except so a hard failure on turbidity
      does not prevent chlorophyll from running (and vice-versa).
    - processing_config is written to session state here so the main() button
      handler can pass it to a resume run later.
    """
    if st.session_state.current_temp_dir is None or not os.path.exists(st.session_state.current_temp_dir):
        st.session_state.current_temp_dir = tempfile.mkdtemp()
    temp_dir = st.session_state.current_temp_dir

    summary_placeholder = st.empty()
    download_summary = dict(st.session_state.download_summary)  # preserve any prior summary

    # ------------------------------------------------------------------
    # Progress bar + stage label — shows overall level (%) and which stage
    # (parameter + month) is currently being processed. Total work units are
    # "months × 2 parameters"; months already recovered via resume/disk cache
    # count as already-done so the bar starts from the right place on Resume.
    # ------------------------------------------------------------------
    start_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    total_months = (end_dt.year - start_dt.year) * 12 + (end_dt.month - start_dt.month)
    total_units = max(total_months * 2, 1)

    progress_bar = st.progress(0)
    stage_text = st.empty()

    def make_progress_callback(stage_label, unit_offset):
        def _callback(done_in_stage, total_in_stage, month_name):
            done_units = unit_offset + done_in_stage
            percent = int(min(1.0, done_units / total_units) * 100)
            if month_name:
                stage_text.markdown(
                    f"**{stage_label}** — در حال پردازش ماه «{month_name}» "
                    f"({done_in_stage} از {total_in_stage}) — {percent}٪"
                )
            else:
                stage_text.markdown(f"**{stage_label}** — در حال آماده‌سازی... — {percent}٪")
            progress_bar.progress(min(1.0, done_units / total_units))
        return _callback

    with st.spinner("در حال پایش کیفیت آب... این فرآیند ممکن است چند دقیقه طول بکشد"):
        # --- Turbidity (NDTI) ---
        turb_results, turb_mean, turb_downloaded, turb_available = [], {}, 0, 0
        try:
            turb_results, turb_mean, turb_downloaded, turb_available = process_single_parameter(
                aoi, start_date, end_date, PARAM_TURBIDITY, temp_dir,
                cloudy_pixel_percentage, scale, resume=resume,
                progress_callback=make_progress_callback("🌊 مرحله ۱ از ۲ — شاخص کدورت (NDTI)", 0)
            )
        except Exception:
            pass  # Partial or zero results; pipeline continues to chlorophyll

        # Merge with any results already in session state (resume case)
        if turb_results:
            st.session_state.results[PARAM_TURBIDITY] = turb_results
            st.session_state.mean_data[PARAM_TURBIDITY] = turb_mean
        download_summary[PARAM_TURBIDITY] = (turb_downloaded, turb_available)

        summary_placeholder.info(
            f"🌊 شاخص کدورت: {turb_downloaded} تصویر از {turb_available} تصویر موجود دریافت شد."
        )

        # --- Chlorophyll-a ---
        chl_results, chl_mean, chl_downloaded, chl_available = [], {}, 0, 0
        try:
            chl_results, chl_mean, chl_downloaded, chl_available = process_single_parameter(
                aoi, start_date, end_date, PARAM_CHLOROPHYLL, temp_dir,
                cloudy_pixel_percentage, scale, resume=resume,
                progress_callback=make_progress_callback("🌿 مرحله ۲ از ۲ — شاخص کلروفیل", total_months)
            )
        except Exception:
            pass  # Partial or zero results; still show whatever was collected

        if chl_results:
            st.session_state.results[PARAM_CHLOROPHYLL] = chl_results
            st.session_state.mean_data[PARAM_CHLOROPHYLL] = chl_mean
        download_summary[PARAM_CHLOROPHYLL] = (chl_downloaded, chl_available)

    progress_bar.progress(1.0)
    stage_text.markdown("✅ پردازش هر دو شاخص به پایان رسید — ۱۰۰٪")

    st.session_state.download_summary = download_summary

    has_any_results = (
        bool(st.session_state.results.get(PARAM_TURBIDITY)) or
        bool(st.session_state.results.get(PARAM_CHLOROPHYLL))
    )
    return has_any_results


# =============================================================================
# Legend + Management Guidance (Persian) — always visible, no jargon
# =============================================================================
def render_turbidity_guidance_panel():
    """Permanently visible legend + management guidance for Turbidity (NDTI)."""
    st.markdown("### 🎨 راهنمای رنگ و تفسیر مدیریتی — شاخص کدورت آب")

    col_legend, col_text = st.columns([1, 2])

    with col_legend:
        fig, ax = plt.subplots(figsize=(5, 0.45))
        cmap = create_turbidity_colormap()
        gradient = np.linspace(0, 1, 256).reshape(1, -1)
        ax.imshow(gradient, aspect='auto', cmap=cmap)
        ax.set_xticks([0, 128, 255])
        ax.set_xticklabels(['آب شفاف', 'متوسط', 'بسیار کدر'])
        ax.set_yticks([])
        st.pyplot(fig)
        plt.close(fig)

    with col_text:
        st.markdown(
            """
**افزایش کدورت آب:**
- کاهش کیفیت آب قابل استفاده برای کشاورزی و مصارف شهری
- افزایش هزینه‌های تصفیه آب
- کاهش نفوذ نور به آب و آسیب به اکوسیستم و آبزیان
- نشانه احتمالی فرسایش خاک، رسوب‌گذاری یا آلودگی در حوضه آبریز

**کاهش کدورت آب:**
- بهبود کیفیت آب و کاهش هزینه‌های تصفیه
- شرایط مطلوب‌تر برای زیست‌بوم آبی و ماهی‌پروری
- نشانه کنترل مؤثر فرسایش و مدیریت بهتر حوضه آبریز

**چرا پایش این شاخص مهم است؟**
پایش روند کدورت به مدیران امکان می‌دهد قبل از وقوع بحران (مانند رسوب‌گذاری در سدها یا
افزایش هزینه تصفیه) اقدام کنند. تغییرات ناگهانی معمولاً نشانه رویدادهایی مانند بارش‌های
شدید، فعالیت‌های عمرانی در بالادست یا تخلیه پساب است و نیازمند بررسی سریع است.
            """
        )


def render_chlorophyll_guidance_panel():
    """Permanently visible legend + management guidance for Chlorophyll-a."""
    st.markdown("### 🎨 راهنمای رنگ و تفسیر مدیریتی — شاخص کلروفیل")

    col_legend, col_text = st.columns([1, 2])

    with col_legend:
        fig, ax = plt.subplots(figsize=(5, 0.45))
        cmap = create_chlorophyll_colormap()
        gradient = np.linspace(0, 1, 256).reshape(1, -1)
        ax.imshow(gradient, aspect='auto', cmap=cmap)
        ax.set_xticks([0, 128, 255])
        ax.set_xticklabels(['کم', 'متوسط', 'بالا (شکوفایی جلبکی)'])
        ax.set_yticks([])
        st.pyplot(fig)
        plt.close(fig)

    with col_text:
        st.markdown(
            """
**افزایش غلظت کلروفیل:**
- احتمال شکوفایی جلبکی و کاهش کیفیت آب آشامیدنی
- افزایش هزینه‌های تصفیه و خطر مسدود شدن فیلترها
- کاهش اکسیژن محلول در آب و خطر برای آبزیان و ماهی‌پروری
- در موارد شدید، احتمال سمیت آب و توقف موقت برداشت آب

**کاهش غلظت کلروفیل:**
- بهبود کیفیت آب و کاهش ریسک‌های بهداشتی
- کاهش هزینه‌های عملیاتی تصفیه‌خانه
- شرایط پایدارتر برای اکوسیستم آبی

**چرا پایش این شاخص مهم است؟**
افزایش ناگهانی کلروفیل معمولاً پیش‌نشانگر شکوفایی جلبکی است که در صورت عدم اقدام به‌موقع
می‌تواند منجر به توقف تأمین آب، هزینه‌های اضطراری تصفیه یا آسیب به صنعت ماهی‌پروری شود.
پایش منظم این شاخص امکان برنامه‌ریزی پیشگیرانه و کاهش ریسک اقتصادی را فراهم می‌کند.
            """
        )


# =============================================================================
# Display: imagery, time-series, and statistics for one parameter
# =============================================================================
def display_side_by_side_imagery(results, parameter_type):
    """Side-by-side processed index image and corresponding RGB image."""
    if not results:
        st.info("داده‌ای برای نمایش در این بازه زمانی وجود ندارد.")
        return

    param_short = "NDTI" if parameter_type == PARAM_TURBIDITY else "Chl-a"

    for r in results:
        if parameter_type == PARAM_TURBIDITY:
            mean_str = f"{r['mean_value']:.4f}" if not np.isnan(r['mean_value']) else "بدون داده"
        else:
            mean_str = f"{r['mean_value']:.2f}" if not np.isnan(r['mean_value']) else "بدون داده"

        cols = st.columns(2)
        cols[0].image(r['wq_image'], caption=f"{r['month_name']} — {param_short}: {mean_str}", use_container_width=True)
        cols[1].image(r['rgb_image'], caption=f"{r['month_name']} — تصویر طبیعی (RGB)", use_container_width=True)


def display_time_series_chart(results, parameter_type):
    """Time series chart of mean index values directly under the imagery."""
    if not results:
        return

    param_short = "NDTI" if parameter_type == PARAM_TURBIDITY else "Chl-a"
    param_unit = "" if parameter_type == PARAM_TURBIDITY else " (µg/L)"
    chart_title = "روند زمانی کدورت آب" if parameter_type == PARAM_TURBIDITY else "روند زمانی کلروفیل"

    months = []
    mean_values = []
    coverage_values = []

    for r in results:
        months.append(r['month_name'])
        mean_values.append(r['mean_value'] if not np.isnan(r['mean_value']) else 0)
        coverage_values.append(r['water_coverage'])

    if not months:
        return

    valid_values = [m for m in mean_values if m != 0]

    fig, ax1 = plt.subplots(figsize=(12, 5))

    color1 = '#1f77b4' if parameter_type == PARAM_TURBIDITY else '#228B22'
    ax1.set_xlabel('Month')
    ax1.set_ylabel(f'Mean {param_short}{param_unit}', color=color1)

    if valid_values:
        ax1.plot(months, mean_values, 'o-', color=color1, linewidth=2, markersize=8, label=f'Mean {param_short}')
        ax1.tick_params(axis='y', labelcolor=color1)

        if parameter_type == PARAM_TURBIDITY:
            ax1.set_ylim(min(mean_values) - 0.02, max(mean_values) + 0.02)
            ax1.axhline(y=0, color='gray', linestyle='--', alpha=0.5, label='Neutral (NDTI=0)')
        else:
            ax1.set_ylim(0, max(mean_values) * 1.2)
    else:
        ax1.text(0.5, 0.5, 'No valid data', ha='center', va='center', transform=ax1.transAxes, fontsize=12)

    ax1.set_xticklabels(months, rotation=45, ha='right')
    ax1.grid(True, alpha=0.3)

    ax1_twin = ax1.twinx()
    color2 = '#2ca02c'
    ax1_twin.set_ylabel('Water Coverage (%)', color=color2)
    ax1_twin.bar(months, coverage_values, alpha=0.3, color=color2, label='Water Coverage')
    ax1_twin.tick_params(axis='y', labelcolor=color2)
    ax1_twin.set_ylim(0, max(coverage_values) * 1.3 if max(coverage_values) > 0 else 100)

    ax1.set_title(chart_title, fontsize=14, fontweight='bold')
    ax1.legend(loc='upper left')

    plt.tight_layout()
    st.pyplot(fig)
    plt.close(fig)


def display_statistics_summary(results, parameter_type):
    """Statistics summary, contained within the same parameter page."""
    if not results:
        return

    param_short = "NDTI" if parameter_type == PARAM_TURBIDITY else "Chl-a"

    months = [r['month_name'] for r in results]
    mean_values = [r['mean_value'] if not np.isnan(r['mean_value']) else 0 for r in results]
    coverage_values = [r['water_coverage'] for r in results]
    valid_values = [m for m in mean_values if m != 0]

    st.markdown("#### 📈 خلاصه آماری")

    col1, col2, col3, col4 = st.columns(4)

    if valid_values:
        if parameter_type == PARAM_TURBIDITY:
            col1.metric(f"میانگین {param_short}", f"{np.mean(valid_values):.4f}")
            col2.metric(f"حداکثر {param_short}", f"{np.max(valid_values):.4f}")
            col3.metric(f"حداقل {param_short}", f"{np.min(valid_values):.4f}")
        else:
            col1.metric(f"میانگین {param_short}", f"{np.mean(valid_values):.2f}")
            col2.metric(f"حداکثر {param_short}", f"{np.max(valid_values):.2f}")
            col3.metric(f"حداقل {param_short}", f"{np.min(valid_values):.2f}")
    else:
        col1.metric(f"میانگین {param_short}", "—")
        col2.metric(f"حداکثر {param_short}", "—")
        col3.metric(f"حداقل {param_short}", "—")

    col4.metric("میانگین پوشش آب", f"{np.mean(coverage_values):.1f}%")

    with st.expander("📋 جدول داده‌های ماهانه"):
        import pandas as pd

        if parameter_type == PARAM_TURBIDITY:
            value_col = [f"{v:.4f}" if v != 0 else "—" for v in mean_values]
        else:
            value_col = [f"{v:.2f}" if v != 0 else "—" for v in mean_values]

        df = pd.DataFrame({
            'ماه': months,
            f'میانگین {param_short}': value_col,
            'پوشش آب (%)': [f"{v:.1f}" for v in coverage_values]
        })
        st.dataframe(df, use_container_width=True)


def generate_combined_timeseries_excel():
    """
    Build a single Excel (.xlsx) workbook containing the monthly time-series
    values for BOTH parameters — Turbidity (NDTI) and Chlorophyll (NDCI) —
    as two sheets in one file. Returns workbook bytes for st.download_button.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
    body_font = Font(name='Arial')
    center = Alignment(horizontal='center')

    center_lat, center_lon = _get_roi_center_coordinates()
    lat_out = round(float(center_lat), 6) if center_lat is not None else "—"
    lon_out = round(float(center_lon), 6) if center_lon is not None else "—"

    sections = [
        (PARAM_TURBIDITY, "کدورت (NDTI)", "میانگین NDTI"),
        (PARAM_CHLOROPHYLL, "کلروفیل (NDCI)", "میانگین NDCI"),
    ]

    for parameter_type, sheet_name, value_header in sections:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.rightToLeft = True

        headers = ["ماه", value_header, "پوشش آب (%)", "عرض جغرافیایی مرکز", "طول جغرافیایی مرکز"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        results = st.session_state.results.get(parameter_type, [])
        for r in sorted(results, key=lambda x: x['month_name']):
            mean_val = r['mean_value']
            mean_out = round(float(mean_val), 4) if not np.isnan(mean_val) else "بدون داده"
            ws.append([r['month_name'], mean_out, round(float(r['water_coverage']), 1), lat_out, lon_out])

        if not results:
            ws.cell(row=2, column=1, value="داده‌ای موجود نیست.").font = body_font
        else:
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = center

        column_widths = [14, 18, 16, 20, 20]
        for i, w in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _get_roi_center_coordinates():
    """
    Return (lat, lon) of the center (centroid) of the region of interest used
    for the current/most recent monitoring run, based on the polygon stored
    in st.session_state.processing_config. Returns (None, None) if no run has
    been configured yet.
    """
    config = st.session_state.get('processing_config')
    if not config or not config.get('polygon_coords'):
        return None, None
    try:
        centroid = Polygon(config['polygon_coords']).centroid
        return centroid.y, centroid.x
    except Exception:
        return None, None


def render_parameter_page(parameter_type):
    """
    Full page for one parameter, in the required order:
    1. Legend + Management Guidance Panel
    2. Side-by-side imagery
    3. Time-series chart
    4. Statistics Summary
    """
    if parameter_type == PARAM_TURBIDITY:
        render_turbidity_guidance_panel()
    else:
        render_chlorophyll_guidance_panel()

    st.divider()

    results = st.session_state.results.get(parameter_type, [])

    if not results:
        st.info("برای مشاهده نتایج، ابتدا یک منطقه را انتخاب و پایش را اجرا کنید.")
        return

    st.markdown("#### 🖼️ تصاویر پردازش‌شده")
    display_side_by_side_imagery(results, parameter_type)

    st.divider()
    display_time_series_chart(results, parameter_type)

    st.divider()
    display_statistics_summary(results, parameter_type)


# =============================================================================
# نظر متخصص آب — Statistical Analysis Pipeline
# (ported as-is from the separate analysis notebook: Mann-Kendall trend test,
#  MAD-based anomaly detection, seasonal climatology, cross-parameter
#  correlation — only the input source changes: in-memory Excel bytes
#  instead of a file path, so nothing needs to be written to disk on Posit.)
# =============================================================================
def _expert_compute_trend(values, n):
    """Mann-Kendall trend test: standard (always) + seasonal (once >=24 months exist)."""
    import pymannkendall as mk

    result = {}

    if n < 4:
        result["standard"] = {"available": False, "reason": f"Mann-Kendall needs at least ~4 points; only {n} available."}
        result["seasonal"] = {"available": False, "reason": "Not enough data."}
        return result

    mk_res = mk.original_test(values)
    result["standard"] = {
        "method": "Mann-Kendall (non-seasonal)",
        "trend": mk_res.trend,
        "significant": bool(mk_res.h),
        "p_value": float(mk_res.p),
        "tau": float(mk_res.Tau),
        "sen_slope_per_month": float(mk_res.slope),
        "note": ("Does not separate the seasonal cycle from the trend. "
                 "If a strong seasonal pattern exists, prefer the seasonal result below when available.")
    }

    if n >= 24:
        try:
            smk_res = mk.seasonal_test(values, period=12)
            result["seasonal"] = {
                "method": "Seasonal Mann-Kendall (Hirsch-Slack), period=12 months",
                "trend": smk_res.trend,
                "significant": bool(smk_res.h),
                "p_value": float(smk_res.p),
                "sen_slope_per_month": float(smk_res.slope),
                "replicates_per_season": round(n / 12, 1),
                "power_caution": (f"Only ~{n/12:.1f} years of data per calendar month are available. "
                                   "Seasonal Mann-Kendall has low statistical power below ~3-4 years; "
                                   "treat this result as directional, not conclusive.")
            }
        except Exception as e:
            result["seasonal"] = {"available": False, "reason": f"Could not compute: {e}"}
    else:
        result["seasonal"] = {
            "available": False,
            "reason": f"Needs at least 24 months (2 full years) to run; only {n} months available."
        }

    return result


def _expert_detect_anomalies_mad(df, date_col, value_column, threshold=3.5):
    """Median Absolute Deviation based anomaly detection."""
    values = df[value_column]
    median = values.median()
    mad = np.median(np.abs(values - median))

    if mad == 0:
        return []

    modified_z = 0.6745 * (values - median) / mad

    anomalies = []
    for idx, row in df.iterrows():
        if abs(modified_z[idx]) > threshold:
            anomalies.append({
                "date": row[date_col].strftime("%Y-%m"),
                "value": float(row[value_column]),
                "modified_z_score": round(float(modified_z[idx]), 2)
            })
    return anomalies


def _expert_analyze_sheet(df, value_column, date_col=None, water_col=None):
    import pandas as pd

    date_col = date_col or df.columns[0]
    water_col = water_col or df.columns[-1]

    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col])
    df = df.sort_values(date_col).reset_index(drop=True)

    values = df[value_column]
    n = len(values)

    raw_observations = [
        {
            "date": row[date_col].strftime("%Y-%m"),
            "value": float(row[value_column]),
            "water_coverage_pct": float(row[water_col])
        }
        for _, row in df.iterrows()
    ]

    trend = _expert_compute_trend(values.values, n)
    anomalies = _expert_detect_anomalies_mad(df, date_col, value_column)

    df["MonthNumber"] = df[date_col].dt.month
    seasonal_climatology = df.groupby("MonthNumber")[value_column].mean().round(4).to_dict()

    summary = {
        "period": {
            "start": df[date_col].min().strftime("%Y-%m"),
            "end": df[date_col].max().strftime("%Y-%m"),
            "months": int(n)
        },
        "raw_observations": raw_observations,
        "statistics": {
            "mean": float(values.mean()),
            "median": float(values.median()),
            "std": float(values.std()),
            "variance": float(values.var()),
            "min": float(values.min()),
            "max": float(values.max()),
            "range": float(values.max() - values.min())
        },
        "trend": trend,
        "extremes": {
            "minimum": {"date": df.loc[values.idxmin(), date_col].strftime("%Y-%m"), "value": float(values.min())},
            "maximum": {"date": df.loc[values.idxmax(), date_col].strftime("%Y-%m"), "value": float(values.max())}
        },
        "seasonal_climatology": seasonal_climatology,
        "water_coverage": {
            "mean_percent": float(df[water_col].mean()),
            "minimum_percent": float(df[water_col].min()),
            "maximum_percent": float(df[water_col].max())
        },
        "missing_values": int(values.isna().sum()),
        "anomalies": anomalies
    }
    return summary


def _expert_compute_correlation_by_date(df1, value_col1, df2, value_col2, date_col1=None, date_col2=None):
    """Correlate NDTI vs NDCI matched by actual date, not row position."""
    import pandas as pd

    date_col1 = date_col1 or df1.columns[0]
    date_col2 = date_col2 or df2.columns[0]

    d1 = df1[[date_col1, value_col1]].copy()
    d1[date_col1] = pd.to_datetime(d1[date_col1])
    d1 = d1.rename(columns={date_col1: "date", value_col1: "v1"})

    d2 = df2[[date_col2, value_col2]].copy()
    d2[date_col2] = pd.to_datetime(d2[date_col2])
    d2 = d2.rename(columns={date_col2: "date", value_col2: "v2"})

    merged = pd.merge(d1, d2, on="date", how="inner")
    unmatched_1 = set(d1["date"]) - set(merged["date"])
    unmatched_2 = set(d2["date"]) - set(merged["date"])

    return {
        "correlation_ndti_ndci": float(merged["v1"].corr(merged["v2"])),
        "n_matched_dates": int(len(merged)),
        "unmatched_dates_ndti_only": sorted(d.strftime("%Y-%m") for d in unmatched_1),
        "unmatched_dates_ndci_only": sorted(d.strftime("%Y-%m") for d in unmatched_2),
    }


def analyze_water_quality_from_bytes(excel_bytes):
    """
    Same logic as the standalone analyze_water_quality(excel_file) from the
    analysis notebook, adapted to read the in-memory Excel bytes produced by
    generate_combined_timeseries_excel() instead of a file on disk.

    Robustness addition (not present in the original notebook): a month with
    no data is written into the Excel export as the text "بدون داده" rather
    than a number, which would otherwise turn the whole column non-numeric
    and silently drop that parameter from the analysis. Those cells are
    coerced to NaN and excluded here instead.

    Also mirrors the notebook's latest addition: the center coordinates of
    the region of interest (columns "عرض جغرافیایی مرکز" / "طول جغرافیایی
    مرکز" in the exported Excel) are lifted into top-level
    "center_latitude" / "center_longitude" keys of the returned summary, so
    the location-aware chat agent can use them directly (e.g. to fetch
    historical weather for that exact point) without re-parsing the sheet.
    """
    import io
    import pandas as pd

    excel_buffer = io.BytesIO(excel_bytes)
    xls = pd.ExcelFile(excel_buffer)
    sheet_names = xls.sheet_names
    results = {}
    cleaned_frames = {}

    for sheet in sheet_names:
        df = pd.read_excel(excel_buffer, sheet_name=sheet)
        numeric_like_cols = [c for c in df.columns if c != df.columns[0]]
        for c in numeric_like_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        df = df.dropna(subset=numeric_like_cols[:1])  # drop rows with no value for the main indicator

        numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
        if len(numeric_cols) < 2 or df.empty:
            continue
        value_col = numeric_cols[0]
        water_col = "پوشش آب (%)" if "پوشش آب (%)" in df.columns else None
        results[sheet] = _expert_analyze_sheet(df, value_col, water_col=water_col)
        cleaned_frames[sheet] = df

    if len(cleaned_frames) >= 2:
        sheet_a, sheet_b = list(cleaned_frames.keys())[:2]
        df1, df2 = cleaned_frames[sheet_a], cleaned_frames[sheet_b]
        col1 = df1.select_dtypes(include=np.number).columns[0]
        col2 = df2.select_dtypes(include=np.number).columns[0]
        results["relationship"] = _expert_compute_correlation_by_date(df1, col1, df2, col2)

    # --- Lift region center coordinates to the top level (for the chat agent) ---
    if cleaned_frames:
        first_sheet_df = list(cleaned_frames.values())[0]
        lat_col = "عرض جغرافیایی مرکز"
        lon_col = "طول جغرافیایی مرکز"
        if lat_col in first_sheet_df.columns and lon_col in first_sheet_df.columns:
            lat_vals = pd.to_numeric(first_sheet_df[lat_col], errors='coerce').dropna()
            lon_vals = pd.to_numeric(first_sheet_df[lon_col], errors='coerce').dropna()
            if not lat_vals.empty and not lon_vals.empty:
                results["center_latitude"] = float(lat_vals.iloc[0])
                results["center_longitude"] = float(lon_vals.iloc[0])

    return results


def _expert_results_signature():
    """Cheap signature used to detect when monitoring results changed, so the
    JSON summary + chat history for نظر متخصص آب can be refreshed automatically."""
    sig = []
    for p in (PARAM_TURBIDITY, PARAM_CHLOROPHYLL):
        results = st.session_state.results.get(p, [])
        sig.append(tuple(sorted(
            (r['month_name'], None if np.isnan(r['mean_value']) else round(float(r['mean_value']), 6))
            for r in results
        )))
    return tuple(sig)


# =============================================================================
# نظر متخصص آب — LLM Agent Chat (LangGraph ReAct agent, OpenAI-compatible)
# =============================================================================
# NOTE ON CREDENTIALS: this app is hosted on a Posit server that cannot read a
# local .env file, so the API base URL / keys (OpenAI-compatible LLM, Tavily,
# OpenWeatherMap) are all hardcoded below and read directly from this file
# instead of being loaded through python-dotenv or environment variables, per
# explicit request. If this project's git repo is ever shared or made public,
# consider moving these values to Posit Connect's own "Environment Variables"
# panel (Settings → Vars) instead of leaving live keys in source — that still
# avoids the .env problem without exposing the keys in version control.
OPENAI_BASE_URL = "https://api.avalai.ir/v1"
OPENAI_API_KEY = "aa-xLsSw3ad4txKKuwvHY7cPGy1StemeS3xtuChVf9utHKOd3Cr"
EXPERT_CHAT_MODEL = "gpt-5.2"  # change to whichever model your endpoint provides

# Tavily web search key, used by the agent's general-purpose web search tool
# (qualitative/general climate context). Hardcoded directly here — read from
# the code, not from a .env file — because the Posit server this app is
# deployed on cannot read a local .env file.
TAVILY_API_KEY = "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"

# OpenWeatherMap Geocoding API key, used by the agent's reverse-geocoding
# tool (lat/lon -> city/region/country name). Also hardcoded directly here
# for the same Posit-server reason as above.
OPENWEATHER_API_KEY = "5fce9bd0bcda8e2cd43468bf50755c82"


def _build_agent_system_prompt(analysis_json):
    """
    Persian system prompt for the water-quality expert agent. Combines the
    original evidence-based / no-hallucination rules with instructions for
    when to use each of the three available tools (reverse geocoding, web
    search, historical weather). The statistical analysis JSON (including,
    when available, center_latitude / center_longitude of the monitored
    region) is embedded directly in the prompt, exactly as in the previous
    non-agent version — it is NOT exposed as a separate callable tool.
    """
    return f"""شما یک متخصص باتجربه در زمینه کیفیت آب، سنجش‌ازدور ماهواره‌ای (سنتینل-۲)، و اقلیم‌شناسی هستید.

در ادامه، خلاصه‌ی تحلیل آماری سری زمانی شاخص کدورت آب و شاخص کلروفیل یک بدنه‌ی آبی، به‌صورت JSON در
اختیار شما قرار گرفته است. این خلاصه شامل مختصات مرکز منطقه (کلیدهای center_latitude و
center_longitude، در صورت وجود)، نتیجه‌ی آزمون روند من-کندال، ناهنجاری‌های شناسایی‌شده (بر پایه‌ی
انحراف مطلق از میانه)، الگوی فصلی چندساله، آمار توصیفی، و همبستگی بین دو شاخص است.

داده‌های تحلیل:
{analysis_json}

شما به سه ابزار دسترسی دارید:
۱. reverse_geocode: یافتن دقیق نام شهر/منطقه/کشور بر اساس یک مختصات جغرافیایی (عرض و طول جغرافیایی)،
   با استفاده از سرویس Geocoding شرکت OpenWeatherMap. هر زمان که نیاز به شناسایی نام منطقه‌ی مورد
   مطالعه بر اساس center_latitude / center_longitude موجود در داده‌های تحلیل بالا داشتید، ابتدا از
   همین ابزار استفاده کنید — نه حدس زدن و نه جست‌وجوی وب — چون این ابزار مستقیماً و با دقت بالا نام
   مکان را از روی مختصات برمی‌گرداند.
۲. tavily_search (جست‌وجوی وب عمومی): برای اطلاعات کیفی و کلی اقلیمی (نوع اقلیم، طبقه‌بندی کوپن،
   الگوهای فصلی بارش و دما، رویدادهای خاص مانند سیل یا خشک‌سالی در آن منطقه) که به‌صورت عددی در دسترس
   نیست و باید از منابع وب یافت شود. همچنین برای هر پرسش عمومی دیگری که نیاز به اطلاعات به‌روز از وب
   دارد.
۳. get_monthly_weather_stats: ابزار دقیقِ داده‌های هواشناسی تاریخی (بایگانی Open-Meteo)، که برای یک
   مختصات جغرافیایی و یک سال/ماه مشخص، دمای بیشینه/کمینه روزانه، بارش، برف، سرعت باد و همچنین درصد
   روزهای برفی آن ماه را برمی‌گرداند. هرگاه کاربر درباره‌ی مقادیر عددی دقیق یک ماه/سال مشخص (دما، بارش،
   سرعت باد، مقدار برف، درصد روزهای برفی) در منطقه‌ی مورد مطالعه سؤال کرد، از این ابزار استفاده کنید —
   نه جست‌وجوی وب. برای مختصات، از center_latitude / center_longitude موجود در داده‌های تحلیل بالا
   استفاده کنید (در صورت نبودن این مقادیر در داده‌ها، صریحاً به کاربر بگویید که مختصات منطقه در دسترس
   نیست).

دستورالعمل‌های استفاده از ابزارها:
- ابزارها را فقط زمانی فراخوانی کنید که پاسخ سؤال کاربر واقعاً به اطلاعات مکانی یا هواشناسی نیاز دارد
  (برای نمونه: «آیا افزایش کدورت در فلان ماه می‌تواند ناشی از بارش شدید یا ذوب برف باشد؟»، «اقلیم این
  منطقه چگونه است؟»، «نام این منطقه چیست؟»). برای سؤال‌هایی که صرفاً درباره‌ی خودِ شاخص کدورت/کلروفیل و
  روند آن‌هاست و پاسخ در داده‌های تحلیل بالا موجود است، نیازی به فراخوانی هیچ ابزاری نیست.
- هرگز مختصات یا نام منطقه را حدس نزنید. برای شناسایی نام منطقه، همیشه ابتدا از reverse_geocode استفاده
  کنید؛ فقط اگر reverse_geocode نتیجه‌ای نداد یا کاربر اطلاعات کیفی/توصیفی بیشتری خواست، سراغ
  tavily_search بروید.
- ترتیب پیشنهادی هنگام نیاز به تفسیر یک نوسان یا ناهنجاری با کمک اطلاعات مکانی/هواشناسی: ابتدا در صورت
  نامشخص بودن نام منطقه، آن را با reverse_geocode شناسایی کنید؛ سپس داده‌ی عددی دقیق مربوط به بازه‌ی
  زمانی موردنظر را با get_monthly_weather_stats بگیرید؛ و در صورت نیاز، زمینه‌ی کلی اقلیمی یا رویدادهای
  خاص را نیز با tavily_search تکمیل کنید.

دستورالعمل‌های پاسخ‌گویی:
۱. پاسخ خود را در درجه‌ی اول بر پایه‌ی داده‌های JSON بالا و در صورت لزوم نتایج ابزارها بنا کنید و از
   دانش عمومی خود درباره‌ی کیفیت آب، سنجش‌ازدور و علوم محیط‌زیست برای تفسیر و تکمیل پاسخ استفاده کنید.
۲. تفسیر را مبتنی بر شواهد ارائه دهید؛ به‌جای نسبت‌دادن هر نوسان یا ناهنجاری به‌طور پیش‌فرض به «خطای
   حسگر»، ابتدا توضیح‌های محیطی، هیدرولوژیکی و هواشناسی محتمل را در نظر بگیرید (رواناب فصلی، بارندگی،
   ذوب برف، سیل، رسوب‌گذاری، شکوفایی جلبکی و مانند آن).
۳. اگر داده‌ی کافی برای نتیجه‌گیری قطعی وجود ندارد (برای نمونه کمتر از ۲۴ ماه برای آزمون فصلی، یا
   نتایج ابزارها ناکافی/متناقض بود)، این محدودیت را صریح بیان کنید؛ حدس قطعی نزنید و هیچ واقعیتی را از
   خود نسازید.
۴. در پاسخ نهایی، در صورت استفاده از ابزارها، بین «شناسایی منطقه»، «داده‌ی عددی دقیق هواشناسی برای
   بازه‌ی درخواستی» و «زمینه‌ی کلی اقلیمی» تمایز قائل شوید.
۵. اگر پرسش کاربر به زبان فارسی باشد، پاسخ باید کاملاً و فقط به زبان فارسی نوشته شود و از هیچ مخفف یا
   واژه‌ی انگلیسی استفاده نشود (برای نمونه به‌جای NDTI بنویسید «شاخص کدورت آب» و به‌جای NDCI بنویسید
   «شاخص کلروفیل»؛ به‌جای MAD بنویسید «انحراف مطلق از میانه»).
۶. اگر پرسش کاربر به زبان دیگری باشد، به همان زبان پاسخ دهید.
"""


def _get_expert_agent(analysis_json):
    """
    Build a fresh LangGraph ReAct agent with the three tools: OpenWeatherMap
    reverse geocoding, Tavily web search, and Open-Meteo historical weather.
    Rebuilt on every call since the system prompt embeds the current
    analysis JSON, which changes whenever new monitoring results are
    generated (see _expert_results_signature()). Building a ReAct agent is
    cheap (no network calls happen until a tool is actually invoked), so
    recreating it per question keeps the code simple and avoids
    stale-prompt bugs.

    API keys (TAVILY_API_KEY, OPENWEATHER_API_KEY) are hardcoded constants
    read directly from this file rather than from a .env file, since the
    Posit server this app runs on cannot read a local .env.

    Requires: langchain, langchain-openai, langgraph, langchain-tavily
    (pip install langchain langchain-openai langgraph langchain-tavily)
    """
    import calendar
    from langchain.chat_models import init_chat_model
    from langchain_core.tools import tool
    from langchain_tavily import TavilySearch
    from langgraph.prebuilt import create_react_agent

    # TavilySearch reads its key from the environment — set it directly from
    # the hardcoded constant above (no .env involved).
    os.environ["TAVILY_API_KEY"] = TAVILY_API_KEY

    model = init_chat_model(
        model=EXPERT_CHAT_MODEL,
        model_provider="openai",
        base_url=OPENAI_BASE_URL,
        api_key=OPENAI_API_KEY,
        temperature=0.3,
    )

    tavily_tool = TavilySearch(max_results=5, topic="general")

    @tool
    def reverse_geocode(lat: float, lon: float) -> str:
        """Reverse-geocode a latitude/longitude pair into a place name
        (city/town/village, state/province, and country) using the
        OpenWeatherMap Geocoding API. Use this whenever you need to identify
        the name of the region/city/country for a given set of coordinates
        (e.g. the center_latitude / center_longitude of the monitored water
        body) — it is faster and more precise than a general web search for
        this specific purpose."""
        params = {
            "lat": lat,
            "lon": lon,
            "limit": 5,
            "appid": OPENWEATHER_API_KEY,
        }
        r = requests.get("https://api.openweathermap.org/geo/1.0/reverse", params=params, timeout=30)
        r.raise_for_status()
        results = r.json()

        if not results:
            return str({"lat": lat, "lon": lon, "results": [], "note": "No place name found for these coordinates."})

        return str({
            "lat": lat,
            "lon": lon,
            "results": [
                {
                    "name": item.get("name"),
                    "state": item.get("state"),
                    "country": item.get("country"),
                    "local_names": item.get("local_names"),
                }
                for item in results
            ],
        })

    @tool
    def get_monthly_weather_stats(lat: float, lon: float, year: int, month: int) -> str:
         """Get daily historical weather data (max/min temperature, precipitation,
            snowfall, wind speed) for a given latitude/longitude and a specific
            year/month, using the Open-Meteo historical archive API. Also returns
            the number of days with snowfall and the snow-day percentage for that
            month. Use this for any request needing exact numeric weather values
            (e.g. 'snow percentage in 2024-01') rather than web search."""
          try:
              data = _fetch_monthly_weather_cached(lat, lon, year, month)
          except Exception as e:
              return str({"error": f"weather service unavailable or timed out: {e}"})
           
        last_day = calendar.monthrange(year, month)[1]
        params = {
            "latitude": lat,
            "longitude": lon,
            "start_date": f"{year}-{month:02d}-01",
            "end_date": f"{year}-{month:02d}-{last_day}",
            "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,snowfall_sum,wind_speed_10m_max",
            "timezone": "auto",
        }
        r = requests.get("https://archive-api.open-meteo.com/v1/archive", params=params, timeout=30)
        r.raise_for_status()
        data = r.json()

        daily = data.get("daily", {})
        snowfall = daily.get("snowfall_sum", [])
        total_days = len(daily.get("time", []))
        snow_days = sum(1 for s in snowfall if s and s > 0)
        snow_pct = round(100 * snow_days / total_days, 1) if total_days else None

        data["summary"] = {
            "total_days": total_days,
            "snow_days": snow_days,
            "snow_day_percentage": snow_pct,
            "temperature_max_monthly": daily.get("temperature_2m_max"),
            "temperature_min_monthly": daily.get("temperature_2m_min"),
            "precipitation_monthly": daily.get("precipitation_sum"),
            "snowfall_monthly": snowfall,
        }
        return str(data)

    tools = [reverse_geocode, tavily_tool, get_monthly_weather_stats]
    system_prompt = _build_agent_system_prompt(analysis_json)
    return create_react_agent(model, tools, prompt=system_prompt)


def ask_water_quality_expert(question, analysis_json, chat_history):
    """
    Send `question` to the water-quality expert agent, giving it the running
    chat history for context. chat_history is a list of
    {"role": "user"/"assistant", "content": ...} dicts — the same shape
    already used elsewhere in the app for st.session_state.expert_chat_history,
    so no other call site needs to change.
    """
    from langchain_core.messages import HumanMessage, AIMessage

    agent = _get_expert_agent(analysis_json)

    messages = []
    for msg in chat_history:
        if msg["role"] == "user":
            messages.append(HumanMessage(content=msg["content"]))
        else:
            messages.append(AIMessage(content=msg["content"]))
    messages.append(HumanMessage(content=question))

    result = agent.invoke({"messages": messages})
    return result["messages"][-1].content


def _inject_persian_chat_css():
    """
    Right-to-left layout + Persian font for the چت با متخصص tab. Streamlit's
    built-in chat elements (st.chat_message / st.chat_input) are LTR by
    default, which misaligns Persian text; this forces RTL direction and a
    Persian-friendly font stack (falls back gracefully if "B Nazanin" is not
    installed on the viewer's system, since it is not a free web font).
    """
    st.markdown(
        """
        <style>
        [data-testid="stChatMessage"],
        [data-testid="stChatMessage"] p,
        [data-testid="stChatMessage"] li,
        [data-testid="stChatMessage"] div,
        [data-testid="stChatMessage"] span {
            direction: rtl;
            text-align: right;
            font-family: "B Nazanin", "BNazanin", "Vazirmatn", Tahoma, sans-serif;
            font-size: 22px;
            line-height: 1.9;
        }
        [data-testid="stChatInput"] textarea {
            direction: rtl;
            text-align: right;
            font-family: "B Nazanin", "BNazanin", "Vazirmatn", Tahoma, sans-serif;
            font-size: 20px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_expert_chat_tab():
    """
    صفحه «نظر متخصص آب»: به‌صورت خودکار خروجی اکسل پایش را می‌گیرد، پایپ‌لاین
    تحلیل آماری موجود را روی آن اجرا می‌کند، خلاصه JSON تولید می‌کند، و یک
    رابط گفتگو با یک عامل هوشمند (LangGraph ReAct agent) در اختیار کاربر
    قرار می‌دهد. این عامل علاوه بر خلاصه JSON، به سه ابزار نیز دسترسی دارد:
    شناسایی نام منطقه از روی مختصات (reverse geocoding با OpenWeatherMap)،
    جست‌وجوی وب برای زمینه‌ی کلی اقلیمی (Tavily)، و دریافت داده‌های دقیق
    هواشناسی تاریخی (Open-Meteo) برای مختصات مرکز منطقه.
    """
    _inject_persian_chat_css()

    st.header("🧑‍🔬 نظر متخصص آب")

    if 'expert_chat_history' not in st.session_state:
        st.session_state.expert_chat_history = []
    if 'expert_analysis_json' not in st.session_state:
        st.session_state.expert_analysis_json = None
    if 'expert_analysis_signature' not in st.session_state:
        st.session_state.expert_analysis_signature = None

    results_turb = st.session_state.results.get(PARAM_TURBIDITY, [])
    results_chl = st.session_state.results.get(PARAM_CHLOROPHYLL, [])

    if not results_turb and not results_chl:
        st.info("برای استفاده از این بخش، ابتدا پایش را اجرا کنید تا داده‌ای برای تحلیل وجود داشته باشد.")
        return

    signature = _expert_results_signature()
    if st.session_state.expert_analysis_json is None or st.session_state.expert_analysis_signature != signature:
        with st.spinner("در حال تحلیل آماری داده‌های سری زمانی..."):
            try:
                excel_bytes = generate_combined_timeseries_excel()
                analysis = analyze_water_quality_from_bytes(excel_bytes)
                st.session_state.expert_analysis_json = json.dumps(analysis, ensure_ascii=False, indent=2)
                st.session_state.expert_analysis_signature = signature
                st.session_state.expert_chat_history = []  # data changed -> start a fresh conversation
            except Exception as e:
                st.error(f"خطا در تحلیل داده‌ها: {e}")
                return

    with st.expander("📄 خلاصه تحلیل (JSON) ارسال‌شده به متخصص هوش مصنوعی"):
        st.code(st.session_state.expert_analysis_json, language="json")

    if st.button("🗑️ شروع گفتگوی جدید", key="expert_chat_reset"):
        st.session_state.expert_chat_history = []
        st.rerun()

    st.divider()

    for msg in st.session_state.expert_chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    user_question = st.chat_input("سؤال خود را درباره کیفیت آب این منطقه بپرسید...")
    if user_question:
        st.session_state.expert_chat_history.append({"role": "user", "content": user_question})
        with st.chat_message("user"):
            st.markdown(user_question)

        with st.chat_message("assistant"):
            with st.spinner("در حال بررسی توسط متخصص هوش مصنوعی..."):
                try:
                    answer = ask_water_quality_expert(
                        user_question,
                        st.session_state.expert_analysis_json,
                        st.session_state.expert_chat_history[:-1]
                    )
                except Exception as e:
                    answer = f"خطا در ارتباط با عامل هوش مصنوعی: {e}"
                st.markdown(answer)

        st.session_state.expert_chat_history.append({"role": "assistant", "content": answer})


# =============================================================================
# Global App Styling — Persian (B Nazanin) font + RTL text + professional
# color palette for general UI chrome only (buttons, headers, inputs,
# sidebar, tabs, metrics, alerts, dataframes, progress bar). This is purely
# presentational: it does not touch the turbidity/chlorophyll colormaps
# (create_turbidity_colormap / create_chlorophyll_colormap) or any of the
# matplotlib figures used to render scientific results, and it does not
# alter any data-processing, calculation, or workflow logic.
# =============================================================================
def _inject_global_app_css():
    """
    Applies a light, water-themed color palette plus the Persian "B Nazanin"
    font (falling back gracefully to Vazirmatn/Tahoma if it is not installed
    on the viewer's system, since it is not a free web font) and right-to-left
    text alignment to Streamlit's general UI chrome. Scoped to text/UI
    elements only — deliberately does not force RTL on the whole document
    body, so widgets such as the folium map, matplotlib figures, and layout
    columns keep their normal structure and behaviour.
    """
    st.markdown(
        """
        <style>
        /* Reliable Persian web font (loads even when "B Nazanin" is not
           installed locally on the viewer's machine — B Nazanin is still
           tried first via the font-family stack below, this is just a
           good-looking, always-available fallback instead of Tahoma). */
        @import url('https://fonts.googleapis.com/css2?family=Vazirmatn:wght@400;500;600;700;800&display=swap');

        /* =====================================================================
           Palette — deep ocean teal + a warm amber accent for emphasis.
           Used only for general UI chrome (never for the turbidity/chlorophyll
           scientific colormaps, which are generated separately in Python).
           ===================================================================== */
        :root {
            --wq-navy:        #0A3F4A;
            --wq-teal-dark:   #0B6E76;
            --wq-teal:        #0E8E99;
            --wq-teal-light:  #2FC2CE;
            --wq-amber:       #F5A524;
            --wq-amber-dark:  #E08E0B;
            --wq-bg-1:        #EAF7F9;
            --wq-bg-2:        #F7FCFD;
            --wq-card:        #FFFFFF;
            --wq-border:      #CDEBEF;
        }

        /* ---- Persian font (applied to text-bearing UI elements) ---- */
        html, body, [class*="css"],
        .stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown span,
        .stText, .stCaption, label, .stButton > button, .stDownloadButton > button,
        .stTextInput input, .stNumberInput input, .stDateInput input,
        .stSelectbox div, .stTabs, .stAlert, .stAlert p,
        [data-testid="stMetricLabel"], [data-testid="stMetricValue"],
        [data-testid="stMetricDelta"], [data-testid="stDataFrame"],
        .streamlit-expanderHeader, h1, h2, h3, h4, h5, h6 {
            font-family: "B Nazanin", "BNazanin", "Vazirmatn", Tahoma, sans-serif;
        }

        /* ---- RTL alignment for Persian text blocks (scoped, not global) ---- */
        .stMarkdown, .stMarkdown p, .stMarkdown li,
        .stAlert, .stAlert p, .streamlit-expanderHeader,
        h1, h2, h3, h4, h5, h6, .stCaption, label {
            direction: rtl;
            text-align: right;
        }

        /* ---- Larger, easier-to-read body / subsection text ---- */
        .stMarkdown p, .stMarkdown li {
            font-size: 1.12rem;
            line-height: 2;
        }
        .stAlert p, .stAlert div {
            font-size: 1.08rem;
            line-height: 1.9;
        }
        .stCaption, [data-testid="stCaptionContainer"] {
            font-size: 1rem !important;
        }

        /* ---- App background: soft, professional water-inspired gradient ---- */
        .stApp {
            background: linear-gradient(160deg, var(--wq-bg-1) 0%, var(--wq-bg-2) 55%, #FDF7EC 100%);
        }

        /* ---- Main page title ---- */
        h1 {
            color: var(--wq-navy);
            font-weight: 800;
            font-size: 2.1rem;
            background: linear-gradient(90deg, var(--wq-navy) 0%, var(--wq-teal) 60%, var(--wq-teal-light) 100%);
            -webkit-background-clip: text;
            background-clip: text;
            -webkit-text-fill-color: transparent;
            border-bottom: 3px solid var(--wq-teal-light);
            padding-bottom: 0.5rem;
            display: inline-block;
        }

        /* ---- Big section titles (st.header, e.g. "1️⃣ ...", "2️⃣ ...", "3️⃣ ...") ---- */
        h2 {
            color: var(--wq-navy) !important;
            font-weight: 800;
            font-size: 1.85rem;
            line-height: 1.6;
            background: linear-gradient(90deg, #DFF4F6 0%, #F3FBFC 85%);
            border-right: 6px solid var(--wq-amber);
            border-radius: 10px;
            padding: 0.7rem 1.1rem;
            margin: 1.6rem 0 1rem 0;
            box-shadow: 0 2px 8px rgba(10, 63, 74, 0.08);
        }

        /* ---- Sub-titles (st.subheader / "#### " markdown) ---- */
        h3 {
            color: var(--wq-teal-dark);
            font-weight: 700;
            font-size: 1.3rem;
            border-right: 3px solid var(--wq-teal-light);
            padding-right: 0.6rem;
        }

        /* ---- Sidebar ---- */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, var(--wq-navy) 0%, var(--wq-teal-dark) 100%);
        }
        section[data-testid="stSidebar"] * {
            color: #EAF7F9 !important;
        }
        section[data-testid="stSidebar"] .stButton > button {
            background: rgba(255,255,255,0.10);
            color: #EAF7F9;
            border: 1px solid rgba(255,255,255,0.35);
            box-shadow: none;
        }
        section[data-testid="stSidebar"] .stButton > button:hover {
            background: rgba(255,255,255,0.22);
            transform: none;
        }

        /* ---- Buttons (general) ---- */
        .stButton > button, .stDownloadButton > button {
            background: linear-gradient(135deg, var(--wq-teal) 0%, var(--wq-teal-light) 100%);
            color: #ffffff;
            border: none;
            border-radius: 12px;
            font-weight: 700;
            padding: 0.55rem 1.4rem;
            transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
            box-shadow: 0 3px 10px rgba(14, 142, 153, 0.28);
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            background: linear-gradient(135deg, var(--wq-teal-dark) 0%, var(--wq-teal) 100%);
            box-shadow: 0 6px 16px rgba(14, 142, 153, 0.38);
            transform: translateY(-2px);
        }
        .stButton > button:active, .stDownloadButton > button:active {
            transform: translateY(0);
        }
        .stButton > button:disabled {
            background: #D7E1E3;
            color: #8FA3A8;
            box-shadow: none;
            transform: none;
        }

        /* ---- Primary call-to-action button (e.g. "🚀 شروع پایش") ---- */
        .stButton > button[kind="primary"],
        .stButton > button[kind="primaryFormSubmit"],
        [data-testid="baseButton-primary"] {
            background: linear-gradient(135deg, var(--wq-amber) 0%, var(--wq-amber-dark) 100%) !important;
            color: #ffffff !important;
            box-shadow: 0 4px 12px rgba(224, 142, 11, 0.35) !important;
        }
        .stButton > button[kind="primary"]:hover,
        .stButton > button[kind="primaryFormSubmit"]:hover,
        [data-testid="baseButton-primary"]:hover {
            background: linear-gradient(135deg, var(--wq-amber-dark) 0%, #C97A08 100%) !important;
            box-shadow: 0 7px 18px rgba(224, 142, 11, 0.45) !important;
            transform: translateY(-2px);
        }

        /* ---- Field labels (e.g. "از تاریخ", "تا تاریخ (غیرشامل)", "🎯 انتخاب منطقه") ---- */
        [data-testid="stWidgetLabel"] p,
        [data-testid="stWidgetLabel"] label,
        .stDateInput label, .stSelectbox label,
        .stTextInput label, .stNumberInput label {
            font-size: 1.15rem !important;
            font-weight: 700 !important;
            color: var(--wq-navy) !important;
        }

        /* ---- Text / date / select inputs ---- */
        .stTextInput input, .stNumberInput input, .stDateInput input {
            border-radius: 10px !important;
            border: 1px solid var(--wq-border) !important;
            font-size: 1.05rem !important;
        }
        .stTextInput input:focus, .stNumberInput input:focus, .stDateInput input:focus {
            border-color: var(--wq-teal) !important;
            box-shadow: 0 0 0 2px rgba(14, 142, 153, 0.18) !important;
        }
        .stSelectbox > div > div {
            border-radius: 10px !important;
            border-color: var(--wq-border) !important;
        }

        /* ---- Tabs ---- */
        .stTabs [data-baseweb="tab-list"] {
            gap: 6px;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: #E1F1F3;
            border-radius: 10px 10px 0 0;
            color: var(--wq-navy);
            font-weight: 700;
            padding: 0.5rem 1rem;
        }
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, var(--wq-teal) 0%, var(--wq-teal-light) 100%) !important;
            color: #ffffff !important;
        }

        /* ---- Metrics ---- */
        [data-testid="stMetric"] {
            background: var(--wq-card);
            border: 1px solid var(--wq-border);
            border-top: 3px solid var(--wq-teal-light);
            border-radius: 12px;
            padding: 14px;
            box-shadow: 0 2px 8px rgba(10, 63, 74, 0.07);
        }

        /* ---- Alerts / info / success / warning boxes ---- */
        .stAlert {
            border-radius: 12px;
            box-shadow: 0 1px 6px rgba(10, 63, 74, 0.06);
        }

        /* ---- Expanders ---- */
        .streamlit-expanderHeader {
            font-weight: 700;
            color: var(--wq-navy);
        }

        /* ---- Progress bar ---- */
        .stProgress > div > div > div {
            background: linear-gradient(90deg, var(--wq-teal) 0%, var(--wq-amber) 100%);
        }

        /* ---- Dataframes / tables ---- */
        [data-testid="stDataFrame"] {
            border-radius: 10px;
            overflow: hidden;
            border: 1px solid var(--wq-border);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


# =============================================================================
# Main Application
# =============================================================================
def main():
    _inject_global_app_css()

    st.title("🌊 سامانه پایش کیفیت آب")
    st.markdown(
        "پایش خودکار **کدورت آب** و **غلظت کلروفیل** با استفاده از تصاویر ماهواره‌ای "
    )

    # Initialize Earth Engine
    ee_ok, ee_msg = initialize_earth_engine()
    if not ee_ok:
        st.error(ee_msg)
        st.stop()

    # ==========================================================================
    # Cache status (kept minimal — no technical internals)
    # ==========================================================================
    has_cache = bool(st.session_state.results.get(PARAM_TURBIDITY)) or \
                bool(st.session_state.results.get(PARAM_CHLOROPHYLL))

    if has_cache:
        st.sidebar.success("✅ نتایج پایش قبلی موجود است")
    else:
        st.sidebar.info("هنوز پایشی انجام نشده است")

    if st.session_state.processing_in_progress:
        st.sidebar.warning("⏳ در حال پردازش...")

    if st.sidebar.button("🗑️ پاک کردن نتایج", disabled=st.session_state.processing_in_progress):
        st.session_state.downloaded_months = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.month_statuses = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.results = {PARAM_TURBIDITY: [], PARAM_CHLOROPHYLL: []}
        st.session_state.mean_data = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.download_summary = {}
        st.session_state.current_temp_dir = None
        st.session_state.processing_config = None
        st.session_state.processing_complete = False
        st.session_state.processing_in_progress = False
        st.rerun()

    # ==========================================================================
    # 1. Region selection
    # ==========================================================================
    st.header("1️⃣ انتخاب منطقه مورد نظر (بدنه آبی)")

    if not st.session_state.processing_in_progress:
        m = folium.Map(location=[35.6892, 51.3890], zoom_start=8)
        plugins.Draw(export=True, position='topleft', draw_options={
            'polyline': False, 'rectangle': True, 'polygon': True,
            'circle': False, 'marker': False, 'circlemarker': False
        }).add_to(m)
        folium.TileLayer(tiles='https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}',
                        attr='Google', name='Satellite').add_to(m)
        folium.LayerControl().add_to(m)

        map_data = st_folium(m, width=800, height=500)

        if map_data and map_data.get('last_active_drawing'):
            geom = map_data['last_active_drawing'].get('geometry', {})
            if geom.get('type') == 'Polygon':
                st.session_state.last_drawn_polygon = Polygon(geom['coordinates'][0])
                st.success("✅ منطقه انتخاب شد")

        if st.button("💾 ذخیره منطقه"):
            if st.session_state.last_drawn_polygon:
                is_duplicate = False
                for existing in st.session_state.drawn_polygons:
                    if existing.equals(st.session_state.last_drawn_polygon):
                        is_duplicate = True
                        break

                if not is_duplicate:
                    st.session_state.drawn_polygons.append(st.session_state.last_drawn_polygon)
                    st.success("✅ منطقه ذخیره شد!")
                    st.rerun()
                else:
                    st.warning("⚠️ این منطقه قبلاً ذخیره شده است")
            else:
                st.warning("⚠️ ابتدا یک منطقه را روی نقشه رسم کنید")
    else:
        st.info("🔒 نقشه در حین پردازش قفل است")

    if st.session_state.drawn_polygons:
        st.subheader("📍 مناطق ذخیره‌شده")
        for i, p in enumerate(st.session_state.drawn_polygons):
            c1, c2, c3 = st.columns([3, 1, 1])
            centroid = p.centroid
            c1.write(f"**منطقه {i+1}**: ~{p.area * 111 * 111:.2f} کیلومتر مربع")
            c2.write(f"مرکز: ({centroid.y:.4f}, {centroid.x:.4f})")
            if c3.button("🗑️", key=f"del_{i}", disabled=st.session_state.processing_in_progress):
                st.session_state.drawn_polygons.pop(i)
                if st.session_state.selected_region_index >= len(st.session_state.drawn_polygons):
                    st.session_state.selected_region_index = max(0, len(st.session_state.drawn_polygons) - 1)
                st.rerun()

    # ==========================================================================
    # 2. Time period
    # ==========================================================================
    st.header("2️⃣ بازه زمانی")
    c1, c2 = st.columns(2)
    start = c1.date_input("از تاریخ", value=date(2024, 1, 1), disabled=st.session_state.processing_in_progress)
    end = c2.date_input("تا تاریخ (غیرشامل)", value=date(2025, 1, 1), disabled=st.session_state.processing_in_progress)

    if start >= end:
        st.error("بازه تاریخ نامعتبر است")
        st.stop()

    months = (end.year - start.year) * 12 + (end.month - start.month)
    st.info(f"📅 بازه انتخابی: **{months} ماه**")

    # ==========================================================================
    # 3. Run analysis — fully automatic (preprocessing + both indices)
    # ==========================================================================
    st.header("3️⃣ اجرای پایش")

    selected_polygon = None

    if st.session_state.drawn_polygons:
        region_options = []
        for i, p in enumerate(st.session_state.drawn_polygons):
            area = p.area * 111 * 111
            region_options.append(f"منطقه {i+1} (~{area:.2f} کیلومتر مربع)")

        if st.session_state.selected_region_index >= len(st.session_state.drawn_polygons):
            st.session_state.selected_region_index = 0

        selected_idx = st.selectbox(
            "🎯 انتخاب منطقه",
            range(len(region_options)),
            format_func=lambda i: region_options[i],
            index=st.session_state.selected_region_index,
            disabled=st.session_state.processing_in_progress
        )

        st.session_state.selected_region_index = selected_idx
        selected_polygon = st.session_state.drawn_polygons[selected_idx]

    elif st.session_state.last_drawn_polygon is not None:
        selected_polygon = st.session_state.last_drawn_polygon
        st.info("ℹ️ استفاده از منطقه رسم‌شده (ذخیره‌نشده)")
    else:
        st.warning("⚠️ ابتدا یک منطقه را روی نقشه رسم کنید")

    st.caption("پس از اجرا، پیش‌پردازش (حذف ابر، حذف برف، استخراج بدنه آب)، سپس شاخص کدورت و شاخص کلروفیل به‌طور خودکار محاسبه می‌شوند.")

    # --- Buttons: Start (fresh) and Resume (after interruption) ---
    btn_col1, btn_col2 = st.columns(2)

    start_btn = btn_col1.button(
        "🚀 شروع پایش",
        type="primary",
        disabled=st.session_state.processing_in_progress or selected_polygon is None
    )

    # Show Resume button only when a previous interrupted run exists
    has_partial_cache = (
        bool(st.session_state.downloaded_months.get(PARAM_TURBIDITY)) or
        bool(st.session_state.downloaded_months.get(PARAM_CHLOROPHYLL)) or
        bool(st.session_state.month_statuses.get(PARAM_TURBIDITY)) or
        bool(st.session_state.month_statuses.get(PARAM_CHLOROPHYLL))
    )
    resume_btn = btn_col2.button(
        "🔄 ادامه از محل قطع",
        disabled=(
            not has_partial_cache or
            st.session_state.processing_config is None or
            st.session_state.processing_in_progress
        ),
        help="اگر اتصال اینترنت قطع شد، پس از اتصال مجدد این دکمه را فشار دهید تا دانلود از همانجا ادامه یابد."
    )

    # --- Auto-continue after an interruption --------------------------------
    # A dropped connection during download/processing does not always surface
    # as a catchable Python exception (e.g. the browser/server connection is
    # cut and the script run is aborted before it reaches its `finally`
    # block). In that case processing_in_progress is left True and survives
    # into the next rerun. We detect that here and resume automatically using
    # the saved processing_config, exactly like pressing "Resume" ourselves.
    # This is the same recovery strategy the old version used (it re-entered
    # processing automatically whenever processing_in_progress was still True
    # on a fresh script run) and is what lets a single click survive an
    # internet interruption instead of leaving the app stuck (Start disabled
    # because processing_in_progress is True, Resume disabled for the same
    # reason).
    auto_continue = (
        not start_btn and not resume_btn
        and st.session_state.processing_in_progress
        and st.session_state.processing_config is not None
    )

    # --- Fresh start ---
    if start_btn:
        st.session_state.downloaded_months = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.month_statuses = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.results = {PARAM_TURBIDITY: [], PARAM_CHLOROPHYLL: []}
        st.session_state.mean_data = {PARAM_TURBIDITY: {}, PARAM_CHLOROPHYLL: {}}
        st.session_state.download_summary = {}
        st.session_state.current_temp_dir = None
        st.session_state.processing_complete = False
        st.session_state.processing_in_progress = True
        st.session_state.resume_after_interruption = False

        # FIX E: Persist processing config so Resume can reconstruct the AOI and params
        st.session_state.processing_config = {
            'polygon_coords': list(selected_polygon.exterior.coords),
            'start_date': start.strftime('%Y-%m-%d'),
            'end_date': end.strftime('%Y-%m-%d'),
            'cloudy_pixel_percentage': CLOUD_THRESHOLD,
            'scale': 10,
        }

        aoi = ee.Geometry.Polygon([list(selected_polygon.exterior.coords)])

        try:
            success = run_full_analysis(
                aoi,
                start.strftime('%Y-%m-%d'),
                end.strftime('%Y-%m-%d'),
                CLOUD_THRESHOLD,
                10,
                resume=False
            )
            st.session_state.processing_complete = success
            if not success:
                st.warning("⚠️ داده‌ای برای این منطقه و بازه زمانی یافت نشد.")
        except Exception:
            # FIX F: On error, flag that a resume is possible instead of losing progress
            st.session_state.resume_after_interruption = True
            st.error(
                "متأسفانه اتصال قطع شد یا خطایی رخ داد. "
                "پس از برقراری اتصال، دکمه «ادامه از محل قطع» را فشار دهید."
            )
        finally:
            st.session_state.processing_in_progress = False
            st.rerun()

    # --- Resume after interruption (manual click or automatic) ---
    if (resume_btn or auto_continue) and st.session_state.processing_config is not None:
        config = st.session_state.processing_config
        st.session_state.processing_in_progress = True
        st.session_state.resume_after_interruption = False

        if auto_continue:
            st.info("🔄 اتصال اینترنت قطع شده بود؛ پایش از همان جا به‌طور خودکار ادامه می‌یابد...")

        aoi = ee.Geometry.Polygon([config['polygon_coords']])

        try:
            success = run_full_analysis(
                aoi,
                config['start_date'],
                config['end_date'],
                config.get('cloudy_pixel_percentage', CLOUD_THRESHOLD),
                config.get('scale', 10),
                resume=True   # FIX G: pass resume=True to skip already-cached months
            )
            # Merge with previously completed results still in session state
            has_any = (
                bool(st.session_state.results.get(PARAM_TURBIDITY)) or
                bool(st.session_state.results.get(PARAM_CHLOROPHYLL))
            )
            if has_any:
                st.session_state.processing_complete = True
            if not success and not has_any:
                st.warning("⚠️ داده‌ای برای این منطقه و بازه زمانی یافت نشد.")
        except Exception:
            st.session_state.resume_after_interruption = True
            st.error(
                "اتصال مجدداً قطع شد. لطفاً دوباره تلاش کنید."
            )
        finally:
            st.session_state.processing_in_progress = False
            st.rerun()

    # Hint when a partial run can be resumed
    if st.session_state.resume_after_interruption and not st.session_state.processing_in_progress:
        st.warning(
            "⚠️ پایش به دلیل قطعی اینترنت متوقف شد. "
            "پس از اتصال مجدد، دکمه «ادامه از محل قطع» را فشار دهید."
        )

    # Simple, user-friendly download summary (persists after run)
    if st.session_state.download_summary:
        st.divider()
        turb_d, turb_a = st.session_state.download_summary.get(PARAM_TURBIDITY, (0, 0))
        chl_d, chl_a = st.session_state.download_summary.get(PARAM_CHLOROPHYLL, (0, 0))
        st.info(
            f"🌊 شاخص کدورت: {turb_d} تصویر از {turb_a} تصویر موجود دریافت شد.\n\n"
            f"🌿 شاخص کلروفیل: {chl_d} تصویر از {chl_a} تصویر موجود دریافت شد."
        )

    # ==========================================================================
    # Results — two dedicated tabs
    # ==========================================================================
    if st.session_state.processing_complete:
        st.divider()
        st.header("📊 نتایج پایش")

        # --- Download combined time-series (Turbidity + Chlorophyll) as one .xlsx ---
        if st.session_state.results.get(PARAM_TURBIDITY) or st.session_state.results.get(PARAM_CHLOROPHYLL):
            st.download_button(
                label="⬇️ دانلود سری زمانی کدورت (NDTI) و کلروفیل (NDCI) — یک فایل Excel",
                data=generate_combined_timeseries_excel(),
                file_name="water_quality_timeseries.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        tab_turbidity, tab_chlorophyll, tab_expert = st.tabs(
            ["🌊 کدورت آب (NDTI)", "🌿 کلروفیل", "💬 چت با متخصص"]
        )

        with tab_turbidity:
            render_parameter_page(PARAM_TURBIDITY)

        with tab_chlorophyll:
            render_parameter_page(PARAM_CHLOROPHYLL)

        with tab_expert:
            render_expert_chat_tab()


if __name__ == "__main__":
    main()
